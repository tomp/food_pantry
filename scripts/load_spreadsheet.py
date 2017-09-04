#!/usr/bin/env python
#
#  Read and work with OLS Caritas 2014-2015 directory
#
import os
import sys
import re
from datetime import datetime, timedelta
import argparse
import logging

logging.basicConfig(format="%(message)s")
log = logging.getLogger(__name__)
log.setLevel(logging.DEBUG)

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

import django
django.setup()
from clients.models import Dependent, Client, Visit

# Constants
CURRENT_SHEET = 'New Data'
NEW_ID_COL = 'New ID #'
NOTES_COL = 'Notes'
AGE7_COL = '#7 age'
FIRST_VISIT = datetime(2017, 1, 7)
FIRST_VISIT_IDX = 32

# Globals
dry_run = False

#######################################################################
# Classes
#######################################################################




#######################################################################
# Main code
#######################################################################

def load_spreadsheet(infile, dryrun=False):
    """Load the client data from the specified spreadsheet.
    If dryrun is True, don't update the django database.
    """
    wb = load_workbook(infile)
    log.info("Opened '{}'".format(infile))

    sheet_names = wb.get_sheet_names()
    log.debug("Found {} sheets:".format(len(sheet_names)))
    for name in sheet_names:
        log.debug("- '{}'".format(name))

    ws = wb[CURRENT_SHEET]
    log.debug("Loaded '{}' sheet".format(CURRENT_SHEET))

    log.debug("Rows: {} - {}\tColumns: {} - {}".format(
        ws.min_row, ws.max_row, ws.min_column, ws.max_column))

    # Check that expected columns exist
    col_names =  [c[0].value for c in ws.iter_cols(min_row=1, max_row=1)]
    for idx, name in enumerate(col_names):
        log.debug("column {:2d}: '{}'".format(idx, name))
    assert NEW_ID_COL in col_names
    # assert NOTES_COL in col_names
    assert AGE7_COL in col_names
    newid_idx = col_names.index(NEW_ID_COL) 
    idNotes_idx = newid_idx + 1
    distNotes_idx = col_names.index(AGE7_COL) + 1
    visit_cols = []
    for idx in range(FIRST_VISIT_IDX, 500):
        name = col_names[idx]
        if not (isinstance(name, datetime) or name.startswith('=')): 
            break
        visit_date = FIRST_VISIT + timedelta(7*len(visit_cols))
        visit_cols.append((idx, visit_date))
        print("col {}: {}".format(idx, visit_date.strftime("%m/%d/%Y")))

    # Load client data, row by row
    for row in ws.iter_rows(min_row=2):
        last_name, first_name, gender, dob, address, city = [
                c.value for c in row[:6]]
        if not (first_name and last_name):
            break

        newId = row[newid_idx].value
        idNotes = row[idNotes_idx].value
        distNotes = row[distNotes_idx].value

        client = Client( last_name=last_name, first_name=first_name,
                gender=gender, dob=dob, address=address, city=city,
                newId=newId, notes=distNotes, reg_notes=idNotes)
        client.save()

        dependents = []
        for idx in range(6, 27, 3):
            name, gender, dob = [c.value for c in row[idx:idx+3]]
            if not name:
                continue
            if ',' in name:
                last_name, first_name = name.split(', ', 1)
            else:
                first_name, last_name = name, ""
            dependent = Dependent(last_name=last_name, first_name=first_name,
                    gender=gender, dob=dob, dependent_on=client)
            dependents.append(dependent)
            dependent.save()

        visits = []
        for idx, visit_date in visit_cols:
            if row[idx].value:
                assert row[idx].value.lower() == 'x'
                visit = Visit(client=client, visited_at=visit_date)
                visits.append(visit)
                visit.save()

        log.debug("Loaded client {}:\t{}\t{} dependents\t{} visits".format(
            client.id, client, len(dependents), len(visits)))


def reset_database():
    Dependent.objects.all().delete()
    Client.objects.all().delete()
    Visit.objects.all().delete()

def parse_args(args):
    """
    Parse the commandline options.
    A namespace object is returned.
    """
    parser = argparse.ArgumentParser(
            "Load food pantry data from Rebecca's Excel spreadsheet")
    parser.add_argument("--infile", "-i",
            help="An Excel spreadsheet with client data")
    parser.add_argument("--reset", action='store_true',
            help="An Excel spreadsheet with client data")
    parser.add_argument("--dry-run", dest="dryrun", action="store_true",
            help="Process the spreadsheet, but don't populate the database.")
    opt = parser.parse_args()
    return opt

def main(args):
    opt = parse_args(args)

    if opt.reset:
        reset_database

    load_spreadsheet(opt.infile)


if __name__ == '__main__':
    import pdb
    try:
        main(sys.argv[1:])
    except Exception:
        # pdb.post_mortem()
        raise
